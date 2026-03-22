#!/usr/bin/env python3
"""
SkillPlus — Excel -> JSON exporter for schedule.

Reads:  data/schedule.xlsx  (sheet: "SCHEDULE")
Writes: assets/data/schedule.flat.json

Optional:
  --grouped  also writes assets/data/schedule.grouped.json
             (grouped by day -> branch -> lessons)

✅ Excel "Day" can be entered in Russian (Пн., ПН, Понедельник и т.п.) or in English codes (mon/tue/...).
✅ Output JSON keeps day as code (mon..sun) + dayLabel in Russian (for table headers/UI).

Usage (from project root):
  python tools/excel_to_json.py
  python tools/excel_to_json.py --excel data/schedule.xlsx --out assets/data/schedule.flat.json --grouped
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


DAY_ORDER = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
DAY_RU = {
    "mon": "Понедельник",
    "tue": "Вторник",
    "wed": "Среда",
    "thu": "Четверг",
    "fri": "Пятница",
    "sat": "Суббота",
    "sun": "Воскресенье",
}



DAY_SHORT_RU = {
    "mon": "Пн.",
    "tue": "Вт.",
    "wed": "Ср.",
    "thu": "Чт.",
    "fri": "Пт.",
    "sat": "Сб.",
    "sun": "Вс.",
}


# Accepted inputs (case-insensitive, trimmed)
DAY_INPUT_MAP = {
    # Monday
    "понедельник": "mon", "пн": "mon", "пон": "mon",
    "monday": "mon", "mon": "mon",
    # Tuesday
    "вторник": "tue", "вт": "tue", "вто": "tue",
    "tuesday": "tue", "tue": "tue", "tues": "tue",
    # Wednesday
    "среда": "wed", "ср": "wed", "сре": "wed",
    "wednesday": "wed", "wed": "wed",
    # Thursday
    "четверг": "thu", "чт": "thu", "чет": "thu",
    "thursday": "thu", "thu": "thu", "thur": "thu", "thurs": "thu",
    # Friday
    "пятница": "fri", "пт": "fri", "пят": "fri",
    "friday": "fri", "fri": "fri",
    # Saturday
    "суббота": "sat", "сб": "sat", "суб": "sat",
    "saturday": "sat", "sat": "sat",
    # Sunday
    "воскресенье": "sun", "вс": "sun", "вос": "sun",
    "sunday": "sun", "sun": "sun",
}


RE_TIME = re.compile(r"^\s*(\d{1,2})[:.](\d{2})\s*$")


def norm_str(v: Any) -> str:
    return str(v).strip()


def parse_day(v: Any) -> Tuple[str, str]:
    """
    Returns (day_code, day_ru_label).
    Raises ValueError if unknown.
    """
    s = norm_str(v).lower()
    s = s.replace("ё", "е")
    s = s.replace(".", "")
    if not s:
        raise ValueError("Day is empty")
    code = DAY_INPUT_MAP.get(s)
    if not code:
        # Try if it's already a code
        if s in DAY_RU:
            code = s
        else:
            raise ValueError(f"Unknown day: {v!r}")
    return code, DAY_RU[code]


def parse_time(v: Any) -> str:
    """
    Supports:
      - '18:00' / '18.00'
      - datetime.time / datetime.datetime
      - Excel time stored as float fraction of day
    Returns HH:MM
    """
    if v is None or v == "":
        raise ValueError("Time is empty")

    if isinstance(v, datetime):
        v = v.time()
    if isinstance(v, time):
        return f"{v.hour:02d}:{v.minute:02d}"

    if isinstance(v, (int, float)):
        # Excel time as fraction of day
        seconds = int(round(float(v) * 24 * 60 * 60))
        seconds %= 24 * 60 * 60
        hh = seconds // 3600
        mm = (seconds % 3600) // 60
        return f"{hh:02d}:{mm:02d}"

    s = norm_str(v)
    m = RE_TIME.match(s)
    if not m:
        raise ValueError(f"Invalid time format: {v!r}. Expected HH:MM")
    hh = int(m.group(1))
    mm = int(m.group(2))
    if not (0 <= hh <= 23 and 0 <= mm <= 59):
        raise ValueError(f"Invalid time value: {v!r}")
    return f"{hh:02d}:{mm:02d}"


def parse_int(v: Any, field: str, allow_blank: bool = False) -> Optional[int]:
    if v is None or v == "":
        return None if allow_blank else _raise(ValueError(f"{field} is empty"))
    try:
        return int(v)
    except Exception as e:
        raise ValueError(f"{field} must be integer, got {v!r}") from e


def _raise(e: Exception):
    raise e


def truthy_active(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return int(v) == 1
    s = norm_str(v).lower()
    return s in {"1", "true", "да", "y", "yes"}


@dataclass
class Lesson:
    id: str
    city: str
    direction: str
    district: str
    branch: str
    ageFrom: int
    ageTo: Optional[int]
    ageLabel: str
    day: str          # mon..sun
    dayLabel: str     # Russian label
    dayShort: str     # Russian short label
    startTime: str    # HH:MM
    endTime: str      # HH:MM
    note: str

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "city": self.city,
            "direction": self.direction,
            "district": self.district,
            "branch": self.branch,
            "ageFrom": self.ageFrom,
            "ageTo": self.ageTo,
            "ageLabel": self.ageLabel,
            "day": self.day,
            "dayLabel": self.dayLabel,
            "dayShort": self.dayShort,
            "startTime": self.startTime,
            "endTime": self.endTime,
            "note": self.note,
        }


def build_age_label(age_from: int, age_to: Optional[int], explicit: Optional[str]) -> str:
    s = (explicit or "").strip()
    if s:
        return s
    if age_to is None:
        return f"{age_from}+"
    if age_to == age_from:
        return f"{age_from}"
    return f"{age_from}–{age_to}"


def read_schedule(excel_path: Path) -> Tuple[List[Lesson], List[str]]:
    wb = load_workbook(excel_path, data_only=True)
    if "SCHEDULE" not in wb.sheetnames:
        raise ValueError('Excel must contain sheet named "SCHEDULE"')

    ws = wb["SCHEDULE"]

    headers = [c.value for c in ws[1]]
    header_idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = ["City", "Direction", "Day", "StartTime", "EndTime", "AgeFrom", "Branch"]
    for r in required:
        if r not in header_idx:
            raise ValueError(f"Missing required column in SCHEDULE: {r!r}. Found: {headers}")

    lessons: List[Lesson] = []
    warnings: List[str] = []

    for row_num in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row_num, c + 1).value for c in range(len(headers))]
        if all(v is None or v == "" for v in row_vals):
            continue  # skip empty rows

        # Active filter
        active_val = row_vals[header_idx.get("Active", -1)] if "Active" in header_idx else 1
        if not truthy_active(active_val):
            continue

        try:
            raw_id = row_vals[header_idx.get("ID", -1)] if "ID" in header_idx else None
            lesson_id = norm_str(raw_id) if raw_id not in (None, "") else f"ROW-{row_num}"

            city = norm_str(row_vals[header_idx["City"]])
            direction = norm_str(row_vals[header_idx["Direction"]])
            district = norm_str(row_vals[header_idx.get("District", -1)]) if "District" in header_idx else ""
            branch = norm_str(row_vals[header_idx["Branch"]])

            age_from = parse_int(row_vals[header_idx["AgeFrom"]], "AgeFrom", allow_blank=False)  # type: ignore
            age_to = parse_int(row_vals[header_idx.get("AgeTo", -1)] if "AgeTo" in header_idx else None, "AgeTo", allow_blank=True)
            explicit_age_label = norm_str(row_vals[header_idx.get("AgeLabel", -1)]) if "AgeLabel" in header_idx and row_vals[header_idx["AgeLabel"]] is not None else ""

            day_code, day_ru = parse_day(row_vals[header_idx["Day"]])
            start_time = parse_time(row_vals[header_idx["StartTime"]])
            end_time = parse_time(row_vals[header_idx["EndTime"]])

            note = norm_str(row_vals[header_idx.get("Note", -1)]) if "Note" in header_idx and row_vals[header_idx["Note"]] is not None else ""

            age_label = build_age_label(age_from, age_to, explicit_age_label)

            lessons.append(
                Lesson(
                    id=lesson_id,
                    city=city,
                    direction=direction,
                    district=district,
                    branch=branch,
                    ageFrom=age_from,
                    ageTo=age_to,
                    ageLabel=age_label,
                    day=day_code,
                    dayLabel=day_ru,
                    dayShort=DAY_SHORT_RU[day_code],
                    startTime=start_time,
                    endTime=end_time,
                    note=note,
                )
            )
        except Exception as e:
            warnings.append(f"Row {row_num}: {e}")

    # Sort: day order, then start time
    order = {d: i for i, d in enumerate(DAY_ORDER)}
    lessons.sort(key=lambda x: (order.get(x.day, 99), x.branch, x.startTime))
    return lessons, warnings


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def build_grouped(lessons: List[Lesson]) -> Dict[str, Any]:
    days: Dict[str, Dict[str, List[Dict[str, Any]]]] = {d: {} for d in DAY_ORDER}
    for l in lessons:
        day_bucket = days.setdefault(l.day, {})
        branch_bucket = day_bucket.setdefault(l.branch, [])
        branch_bucket.append(l.to_dict())

    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "dayLabels": DAY_RU,
        },
        "days": days,
    }


def main():
    parser = argparse.ArgumentParser(description="Excel -> JSON exporter for schedule")
    parser.add_argument("--excel", default="data/schedule.xlsx", help="Path to Excel file (default: data/schedule.xlsx)")
    parser.add_argument("--out", default="assets/data/schedule.flat.json", help="Output JSON path (default: assets/data/schedule.flat.json)")
    parser.add_argument("--grouped", action="store_true", help="Also write grouped JSON to assets/data/schedule.grouped.json")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    lessons, warnings = read_schedule(excel_path)

    out_path = Path(args.out)
    write_json(out_path, [l.to_dict() for l in lessons])
    print(f"✅ Wrote {len(lessons)} lessons -> {out_path}")

    if args.grouped:
        grouped_path = out_path.parent / "schedule.grouped.json"
        write_json(grouped_path, build_grouped(lessons))
        print(f"✅ Also wrote grouped -> {grouped_path}")

    if warnings:
        print("\n⚠️ Warnings (these rows were skipped or had issues):")
        for w in warnings:
            print(" -", w)


if __name__ == "__main__":
    main()
